# from openpyxl import load_workbook

# wb = load_workbook(filename='/Users/fkl/vr/dataset/test.xlsx')
# ws = wb.worksheets[0]
# img = openpyxl.drawing.Image('test.jpg')
# img.anchor(ws.cell('E1'))
# ws.add_image(img)

# img = openpyxl.drawing.Image('test.jpg')
# img.anchor(ws.cell('E2'))
# ws.add_image(img)
# wb.save('out.xlsx')

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

wb = load_workbook(filename='/Users/fkl/vr/dataset/test.xlsx')
ws = wb.worksheets[0]

for i in range(2, 4456):
    link = str(ws['D'+str(i)].value)
    image_name = link.split("/")[-1]
    image_path = "/Users/fkl/vr/dataset/web_low_res/" + image_name
    print(image_path)
    img = Image(image_path)
    ws.add_image(img, 'E'+str(i))


# ws.add_image(img1)
# ws.add_image(img2)

#ws.add_image(img1, 'B2')
#ws.add_image(img2, 'B14')

wb.save('excel-image.xlsx')
